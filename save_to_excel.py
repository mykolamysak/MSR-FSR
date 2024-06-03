import openpyxl
from openpyxl.drawing.image import Image


def save_to_excel(self):
    try:
        workbook = openpyxl.load_workbook('results.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    worksheet = workbook.active

    # Find the last filled row
    last_row = worksheet.max_row

    # Write values to the next row
    row = last_row + 1
    worksheet.cell(row=row, column=1, value=self.degree_A_combobox.get())
    worksheet.cell(row=row, column=2, value=self.degree_B_combobox.get())
    worksheet.cell(row=row, column=3, value=self.selected_i)
    worksheet.cell(row=row, column=4, value=self.selected_j)
    worksheet.cell(row=row, column=5, value=self.rank_combobox.get())
    worksheet.cell(row=row, column=6, value=self.period_label['text'].split(':')[1].strip())
    worksheet.cell(row=row, column=7, value=self.period_exp_label['text'].split(':')[1].strip())
    worksheet.cell(row=row, column=8, value=self.hamming_weight_value['text'].split(':')[1].strip())
    worksheet.cell(row=row, column=9, value=self.hamming_weight_exp_value['text'].split(':')[1].strip())

    # # Save the Matplotlib plot as an image file
    # image_path = "plot.png"
    # self.figure.savefig(image_path)
    #
    # # Insert the image into the Excel file
    # img = Image(image_path)
    # img.width, img.height = 100, 80  # Set image size
    # img.anchor = f'J{row}'  # Insert the image at the 10th column
    #
    # worksheet.add_image(img)

    #Save the file
    workbook.save("results.xlsx")
